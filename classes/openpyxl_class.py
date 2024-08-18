import re
from env.env import NAME
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins


class OpenpyxlHandler:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.font_camb_12 = Font(name="Cambria", size=12, bold=True)

        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.double_bottom_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="double"),
        )

        self.center_align = Alignment(horizontal="center", vertical="bottom")

    def set_data_in_excel(self, data):
        for r in dataframe_to_rows(data, index=False, header=True):
            self.ws.append(r)

    def add_header(self):
        self.ws.oddHeader.left.text = NAME

        self.ws.oddHeader.left.size = 14
        self.ws.oddHeader.left.font = "Cambria,Bold"
        self.ws.oddHeader.center.text = "&[Date]"

        self.ws.oddHeader.center.size = 11
        self.ws.oddHeader.center.font = "Cambria,Bold"
        self.ws.oddHeader.right.text = "PAGE &[Page] OF &[Pages]"

        self.ws.oddHeader.right.size = 11
        self.ws.oddHeader.right.font = "Cambria,Bold"

    def format_heading_and_body_height(self):
        body_font = Font(name="Calibri", size=12)

        # Apply styles
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = self.alignment
                if cell.row == 1:  # Header row
                    cell.font = self.font_camb_12
                else:  # Body rows
                    cell.font = body_font
                cell.border = self.thin_border

        # Set row heights
        self.ws.row_dimensions[1].height = 49.5

        for row in self.ws.iter_rows(min_row=2):
            self.ws.row_dimensions[row[0].row].height = (
                25  # Adjust the height as needed
            )

    def format_width(self):
        # Set column widths # Set width for 'Client (Full Name)' column
        self.ws.column_dimensions["A"].width = 60
        self.ws.column_dimensions["B"].width = 37.14
        self.ws.column_dimensions["C"].width = 13.43
        self.ws.column_dimensions["D"].width = 12.14
        self.ws.column_dimensions["E"].width = 12.14
        self.ws.column_dimensions["F"].width = 12.14
        self.ws.column_dimensions["G"].width = 12.14
        self.ws.column_dimensions["H"].width = 11.14
        self.ws.column_dimensions["I"].width = 13.43
        self.ws.column_dimensions["J"].width = 13.14

    def set_page_margin(self):
        cm_to_inch = 1 / 2.54

        self.ws.page_margins = PageMargins(
            left=1.1 * cm_to_inch,
            right=1.1 * cm_to_inch,
            top=1.9 * cm_to_inch,
            bottom=1.9 * cm_to_inch,
            header=0.8 * cm_to_inch,
            footer=0.8 * cm_to_inch,
        )

    def set_print_options(self):

        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE

        self.ws.print_options.horizontalCentered = True
        self.ws.print_options.verticalCentered = True
        self.ws.page_setup.scale = 94

    def add_formula_and_format_currency(self):
        for row in range(2, self.ws.max_row + 1):
            self.ws[f"I{row}"] = f"=C{row}*H{row}*0.65"

        for row in self.ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = "[$$-409]#,##0.00"

        for row in self.ws.iter_rows(min_row=2, min_col=9, max_col=9):
            for cell in row:
                cell.number_format = "[$$-409]#,##0.00"

        for row in self.ws.iter_rows(
            min_row=2, min_col=8, max_col=8
        ):  # Column H is the 8th column
            for cell in row:
                cell.number_format = "0.00"

    def add_format_total_cell(self):
        self.last_row = self.ws.max_row + 1

        # Merge cells from column B to column G in the last row
        self.ws.merge_cells(
            start_row=self.last_row, start_column=2, end_row=self.last_row, end_column=7
        )

        # Write "TOTAL:" in the merged cell
        cell = self.ws.cell(row=self.last_row, column=2, value="TOTAL:")

        # Set the font to Cambria, size 12, bold
        cell.font = self.font_camb_12

        # Align the text to the bottom and right
        cell.alignment = Alignment(horizontal="right", vertical="bottom")

        for col in range(1, 11):  # Columns A to G
            self.ws.cell(row=self.last_row, column=col).border = self.thin_border

        self.ws.row_dimensions[self.last_row].height = 25

    def add_format_total_sum_cell(self):
        sum_formula = f"=SUM(H{2}:H{self.last_row - 1})"
        sum_cell = self.ws.cell(row=self.last_row, column=8, value=sum_formula)

        # Set the font to Cambria, size 12, bold
        sum_cell.font = self.font_camb_12

        # Align the text to the bottom and center
        sum_cell.alignment = self.center_align

        # Add borders to the cell with a double bottom border

        sum_cell.border = self.double_bottom_border
        sum_cell.number_format = "0.00"

    def add_format_amount_sum_cell(self):
        sum_formula_I = f"=SUM(I{2}:I{self.last_row - 1})"
        sum_cell_I = self.ws.cell(row=self.last_row, column=9, value=sum_formula_I)

        # Set the font to Cambria, size 12, bold
        sum_cell_I.font = self.font_camb_12

        # Align the text to the bottom and center
        sum_cell_I.alignment = self.center_align

        # Add borders to the cell with a double bottom border
        sum_cell_I.border = self.double_bottom_border

        sum_cell_I.number_format = "$#,##0.00"

    def add_format_declaration_cell(self):
        self.last_row += 2

        # Merge cells from column B to column F in the last row
        self.ws.merge_cells(
            start_row=self.last_row, start_column=2, end_row=self.last_row, end_column=6
        )

        # Write the specified text in the merged cell
        cell = self.ws.cell(
            row=self.last_row,
            column=2,
            value="I agree the information above is accurate to the best of my knowledge.",
        )

        # Set the font to Cambria, size 12, bold and italic
        cell.font = Font(name="Cambria", size=12, bold=True, italic=True)

        # Align the text to the middle and center
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set the height of the last row to 24.5
        self.ws.row_dimensions[self.last_row].height = 24.5

    def add_format_signature_cell(self):

        self.ws.merge_cells(
            start_row=self.last_row,
            start_column=7,
            end_row=self.last_row,
            end_column=10,
        )

        # Write "Copilot is Awesome" in the merged cell
        cell = self.ws.cell(row=self.last_row, column=7, value=NAME)

        # Set the font to Cambria, size 12, bold
        cell.font = Font(name="Cambria", size=12, bold=True)

        # Align the text to the middle and center
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add a thick bottom border to the cell
        thick_bottom_border = Border(bottom=Side(style="thick"))

        for col in range(7, 11):  # Columns G to J
            self.ws.cell(row=self.last_row, column=col).border = thick_bottom_border

    def highlight_anomalous_hours(self):
        "Check for 'hrs' and update Column H"
        # Define the yellow fill
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        # Iterate through the rows in Column A
        for row in self.ws.iter_rows(min_row=2, max_col=1, max_row=self.last_row - 3):
            cell = row[0]
            if "hrs" in str(cell.value):
                hours = re.search(r"(\d+)\s*hrs", cell.value)
                if hours:
                    self.ws.cell(row=cell.row, column=8, value=int(hours.group(1)))

            if "min" not in str(cell.value):
                # Highlight the entire row in yellow
                for cell in row:
                    for col in self.ws.iter_cols(
                        min_col=1,
                        max_col=self.ws.max_column,
                        min_row=cell.row,
                        max_row=cell.row,
                    ):
                        for cell in col:
                            cell.fill = yellow_fill

    def highlight_duplicate_cells(self):
        pattern = re.compile(
            r"(?:(?:Mrs\.|Miss|Mr\.?|Ms\.?)\s)?([A-Za-z]+(?:\s[A-Za-z]+)?)\s([A-Za-z']+)"
        )
        # Define the fill color for highlighting
        grey_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        # Dictionary to store names and their row numbers
        name_dict = {}
        # Iterate through cells in column A
        for row in range(1, self.ws.max_row + 1):
            cell_value = self.ws.cell(row=row, column=1).value
            if cell_value:
                match = pattern.match(cell_value)
                if match:
                    full_name = f"{match.group(1)} {match.group(2)}"
                    if full_name in name_dict:
                        # Highlight both the current row and the previously found row
                        for col in range(1, self.ws.max_column + 1):
                            self.ws.cell(row=row, column=col).fill = grey_fill
                            self.ws.cell(row=name_dict[full_name], column=col).fill = (
                                grey_fill
                            )
                    else:
                        name_dict[full_name] = row

    def save_to_excel(self, filename):
        self.wb.save(filename)
